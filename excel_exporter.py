"""
excel_exporter.py
Converts parsed JSON contract data into a formatted Excel (.xlsx) file.
"""

import json
import os

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill

from config import OUTPUT_FOLDER, CLAUSE_CHAR_LIMIT


# ── Column layout ──────────────────────────────────────────────────────────────

COLUMN_WIDTHS = {
    "A": 30, "B": 60, "C": 40, "D": 20, "E": 20,
    "F": 20, "G": 20, "H": 20, "I": 30, "J": 90, "K": 90,
}

WRAP_COLUMNS  = {"J", "K"}      # top-aligned, text-wrapped
HEADER_COLOR  = "78A2CC"
HEADER_HEIGHT = 30
ROW_HEIGHT    = 140


# ── Public function ────────────────────────────────────────────────────────────

def export(data: list[dict], output_name: str) -> str:
    """
    Split oversized Clause Text into Notes, flatten to DataFrame, write Excel.
    Returns the path of the saved file.
    """
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    processed = _split_clause_notes(data)
    output_path = os.path.join(OUTPUT_FOLDER, f"formatted_{output_name}.xlsx")

    _write_excel(processed, output_path)
    return output_path


# ── Internal helpers ───────────────────────────────────────────────────────────

def _split_clause_notes(data: list[dict]) -> list[dict]:
    """Move Clause Text that exceeds CLAUSE_CHAR_LIMIT overflow into Notes."""
    result = []
    for item in data:
        clauses = item.get("Clause Text", [])
        if isinstance(clauses, str):
            clauses = [clauses]

        kept, overflow = [], ""
        total = 0
        for line in clauses:
            if total + len(line) <= CLAUSE_CHAR_LIMIT:
                kept.append(line)
                total += len(line)
            else:
                overflow += line + " "

        item = dict(item)   # don't mutate original
        item["Clause Text"] = "\n".join(c for c in kept if c != "]")
        item["Notes"]       = overflow.strip()
        result.append(item)
    return result


def _write_excel(data: list[dict], path: str) -> None:
    """Write data to a formatted Excel workbook."""
    df = pd.DataFrame(data)
    df.to_excel(path, index=False)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Column widths
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Header row styling
    ws.row_dimensions[1].height = HEADER_HEIGHT
    fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill

    # Data row heights + alignment
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column_letter in WRAP_COLUMNS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)
